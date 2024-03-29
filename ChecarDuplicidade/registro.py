import csv
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from random import randint
from datetime import date
from classes import DadosCliente, DadosProduto
import os


def cria_clientes(micros: set, plan: str) -> list[DadosCliente]:
    """
    Função que recebe as regiões e a planilha e retorna uma lista de objetos do tipo DadosCliente.
    Essa função garante a rotatividade de clientes por região, selecionando aleatoriamente.
    :param micros: um conjunto de strings representando as microrregiões
    :param plan: string indicando o caminho onde se encontra a base de clientes vs. microrregiões
    :return: retorna uma lista de objetos da classe DadosCliente, com dados do cliente
    """
    df = pd.read_excel(
        plan,
        sheet_name="Base_Clientes_Consulta",
        usecols=["MicroRegiao", "CNPJ", "Cod_Cliente"],
        dtype=str,
    )
    df = df.fillna("-")
    # listando todos os clientes
    cods = list(df.itertuples(index=False, name=None))
    # filtrando os clientes por regiao desejada
    cods = list(filter(lambda x: x[0] in micros, cods))
    unico = []
    for micro in micros:
        # escolhendo apenas um cliente por região
        clientes = list(filter(lambda x: x[0] == micro, cods))
        try:
            unico.append(clientes[randint(0, len(clientes) - 1)])
        except ValueError:
            print(micro)
            pass
    return [
        DadosCliente(
            identificacao=un,
            cimento_todas_obras={"25 to": 0, "42 to": 0, "50 to": 0},
            cimento_estrutural={"40 e": 0, "50 e": 0},
            codigos_analisados=[],
        )
        for un in unico
    ]


# Usar a vistoria diaria para garantir ao menos 4 checks de todas as Micros por mês (2000 checks/mês)
def geracao_vistoria_diaria(plan: str, dia_semana: int) -> list[DadosCliente]:
    """
    Gera uma lista de 100 clientes, de micros diferentes. As micros verificadas variam de acordo com o dia da semana
    :param plan: o caminho da planilha com a base de clientes por Micro
    :param dia_semana: inteiro representando o dia da semana (segunda:0 até sexta: 4). Domingo (6) e Sábado (5) não são
    levados em conta.
    :return: Retorna uma lista de objetos da classe DadosClientes, que contém informações sobre o cliente.
    """
    df = pd.read_excel(plan, sheet_name="Consulta", usecols=["Microregiao"])
    df.sort_values(by=["Microregiao"], inplace=True)
    micros = set(df.Microregiao[dia_semana * 102 : (dia_semana + 1) * 102])
    return cria_clientes(micros, plan)


def seleciona_aba(wb: Workbook, aba: str):
    """Função que seleciona a aba desejada"""
    for (i, s) in enumerate(wb.sheetnames):
        if wb.sheetnames[i] == aba:
            wb.active = i
    return wb.active


def report_duplicidade(clientes: list[DadosCliente], plan: str) -> int:
    """Recebe a lista de Dadoscliente e registra na planilha recebida. Usa-se openpyxl para registrar"""
    wb = load_workbook(plan, data_only=True)
    today = date.today()
    sheet = seleciona_aba(wb=wb, aba="Consulta")
    cab = [cell.value for cell in sheet[1]]  # Gerando o cabeçalho
    print(cab)
    counter = sheet.max_row
    qt_dup = 0
    # Checando a microregião e anotando a duplicidade.
    for cliente in clientes:
        if not ("comentario" in cliente.cimento_todas_obras.keys()):
            check_to = cliente.cimento_todas_obras
            check_e = cliente.cimento_estrutural
            for i in range(2, int(counter) + 2):
                if sheet[i][cab.index("Microregiao")].value == cliente.identificacao[0]:
                    sheet[i][cab.index("Cod_Cliente")].value = cliente.identificacao[2]
                    sheet[i][cab.index("CNPJ")].value = cliente.identificacao[1]
                    for key, qtd in tuple(check_to.items()):
                        sheet[i][cab.index(key)].value = qtd
                    for key, qtd in tuple(check_e.items()):
                        sheet[i][cab.index(key)].value = qtd
                    try:
                        if (
                            sheet[i][cab.index("50 to")].value > 1
                            or sheet[i][cab.index("42 to")].value > 1
                        ):
                            sheet[i][cab.index("Duplicidade")].value = "sim"
                            qt_dup += 1
                        else:
                            sheet[i][cab.index("Duplicidade")].value = "não"
                    except TypeError:
                        sheet[i][cab.index("Duplicidade")].value = ""

                    sheet[i][cab.index("códigos analisados")].value = ", ".join(
                        cliente.codigos_analisados
                    )
                    sheet[i][cab.index("Data Consulta")].value = today.strftime(
                        "%d/%m/%y"
                    )
    wb.save(plan)
    return qt_dup


def criando_relatorio_exec_int(clientes: list[DadosCliente]):
    """Gera um arquivo .csv e .xlsx com um relatório do que foi encontrado"""
    today = date.today()
    relatorio = os.path.join(
        "C:" + os.sep,
        "Users" + os.sep,
        "lucashl" + os.sep,
        "Desktop" + os.sep,
        "VCimentos" + os.sep,
        "RPA" + os.sep,
        "ChecarDuplicidade" + os.sep,
        "relatorios" + os.sep,
        f"relatorio_exec_int_{today.strftime('%d.%m.%y')}.csv",
    )
    with open(relatorio, "w", encoding="UTF8", newline="") as file:
        writer = csv.writer(file, dialect="excel")
        header = [
            "CNPJ",
            "Cod_Cliente1",
            "Regiao",
            "T.O. 25kg",
            "T.O. 42kg",
            "T.O. 50kg",
            "Est. 40kg",
            "Est. 50kg",
            "códigos analisados",
            "duplicidade",
            "códigos a sanear",
            "comentário E.I.",
        ]
        writer.writerow(header)
        for c in clientes:
            try:
                if (
                    c.cimento_todas_obras["42 to"] > 1
                    or c.cimento_todas_obras["50 to"] > 1
                ):
                    dupli = "sim"
                else:
                    dupli = "não"
                data = [
                    c.identificacao[1],
                    c.identificacao[2],
                    c.identificacao[0],
                    c.cimento_todas_obras["25 to"],
                    c.cimento_todas_obras["42 to"],
                    c.cimento_todas_obras["50 to"],
                    c.cimento_estrutural["40 e"],
                    c.cimento_estrutural["50 e"],
                    ", ".join(c.codigos_analisados),
                    dupli,
                    "",
                    "",
                ]
                writer.writerow(data)
            except KeyError:
                pass
    df = pd.read_csv(relatorio)
    df.to_excel(
        os.path.join(
            os.sep + os.sep + "brcwbwvfs01vc",
            "COMERCIA",
            "Relatorios",
            "50 - CAC",
            "Gestão EO",
            "JS+",
            "Remoção Duplicidades",
            "Análise micro X produto LV",
            "Evidências Duplic",
            "relatorios",
            f"relatorio_exec_int_{today.strftime('%d.%m.%y')}.xlsx",
        )
    )


def enviar_email(qtd: int, emails: tuple, plan: str, dupli: int) -> str:
    today = date.today()
    relatorio = (
        f"Prezados,\nA tarefa de check duplicidade de produtos na Loja virtual em "
        f"{today.strftime('%d.%m.%y')} foi concluída.\nForam checadas {qtd} micro-regiões, das"
        f" quais {dupli} estão com duplicidades. o check está registrado em {plan}\nAtt."
    )
    return relatorio
