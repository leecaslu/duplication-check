# Programa que realiza consulta à lista .xlsx de clientes a serem consultados. Utiliza pandas e openpyxl.
# Programa também realiza registro em arquivos .xlsx, .csv e .txt
"""
TODO Implementar função que manda email com a confirmação de finalização da tarefa.
TODO Garantir rotatividade de clientes semanalmente
"""

# imports necessarios: openpyxl, selenium.

import csv
import pandas as pd
from ast import literal_eval
from openpyxl import load_workbook
from openpyxl import Workbook
from random import randint
from datetime import date
from dados_duplicidade import DadosCliente
from os import makedirs

# variáveis úteis
conv_alfa_num = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't',
                 'u', 'v', 'w', 'x', 'y', 'z']


def seleciona_aba(wb: Workbook, aba: str):
    """Função que seleciona a aba desejada"""
    for (i, s) in enumerate(wb.sheetnames):
        if wb.sheetnames[i] == aba:
            wb.active = i
    return wb.active


def cria_clientes(micros: set, plan: str) -> list[DadosCliente]:
    """Função que recebe as regiões e a planilha e retorna uma lista de objetos do tipo DadosCliente.
    Essa função garante a rotatividade de clientes por região"""
    df = pd.read_excel(plan, sheet_name='Base_Clientes_Consulta', usecols=["Regiao", "CNPJ", "Código"])
    # listando todos os clientes
    cods = list(df.itertuples(index=False, name=None))
    print(cods)
    # filtrando os clientes por regiao desejada
    cods = list(filter(lambda x: x[2] in micros, cods))
    print(cods)
    unico = []
    for micro in micros:
        # escolhendo apenas um cliente por região
        clientes = list(filter(lambda x: x[2] == micro, cods))
        unico.append(clientes[randint(0, len(clientes)-1)])
    return [DadosCliente(identificacao=un, cimento_todas_obras={'25 to': 0, '50 to': 0},
                         cimento_estrutural={'40 e': 0, '50 e': 0}, codigos_analisados=[]) for un in unico]


def geracao_relatorio_mensal(plan: str) -> list[DadosCliente]:
    """Gerando uma lista de clientes, um de cada região"""
    df = pd.read_excel(plan, sheet_name='Consulta', usecols=["Regiao", "Duplicidade"])
    micros = set(df.Regiao)
    print(micros)
    return cria_clientes(micros, plan)


def geracao_vistoria_diaria(plan: str, qtd: int) -> list[DadosCliente]:
    """Gerando uma lista de x clientes, um de cada região, de maneira aleatória. Foca nos clientes com duplicidade"""
    df = pd.read_excel(plan, sheet_name='Consulta', usecols=["Regiao", "Duplicidade"])
    counter = len(df.index)
    micros = set()
    while len(micros) < qtd-5:
        i = randint(0, counter-1)
        if df.Duplicidade[i] == 'Sim':
            micros.add(df.Regiao[i])
    while len(micros) < qtd:
        i = randint(0, counter-1)
        if df.Duplicidade[i] == 'Não':
            micros.add(df.Regiao[i])
    print(micros)
    return cria_clientes(micros, plan)


def geracao_vistoria_filtrada(plan: str, filtro: str) -> list:
    df = pd.read_excel(plan, sheet_name='Consulta', usecols=["Regiao", "Duplicidade"])
    counter = len(df.index)
    micros = set()
    for i in range(counter):
        if df.Regiao[i].find(filtro) > 0:
            micros.add(df.Regiao[i])
    return cria_clientes(micros, plan)


def report_duplicidade(clientes: list[DadosCliente], plan: str):
    """Recebe a lista de Dadoscliente e registra na planilha recebida. Usa-se openpyxl para registrar"""
    wb = load_workbook(plan, data_only=True)
    today = date.today()
    sheet = seleciona_aba(wb=wb, aba='Consulta')
    cab = [cell.value for cell in sheet[1]]  # Gerando o cabeçalho
    print(cab)
    counter = sheet.max_row
    qt_dup = 0
    # Checando a microregião e anotando a duplicidade.
    for cliente in clientes:
        check_to = cliente.cimento_todas_obras
        check_e = cliente.cimento_estrutural
        for i in range(2, int(counter) + 2):
            if sheet[i][cab.index('Regiao')].value == cliente.identificacao[2]:
                for key, qtd in tuple(check_to.items()):
                    sheet[i][cab.index(key)].value = qtd
                for key, qtd in tuple(check_e.items()):
                    sheet[i][cab.index(key)].value = qtd
                try:
                    if (sheet[i][cab.index('25 to')].value > 1 or sheet[i][cab.index('50 to')].value > 1
                            or sheet[i][cab.index('40 e')].value > 1 or sheet[i][cab.index('50 e')].value > 1):
                        sheet[i][cab.index('Duplicidade')].value = 'sim'
                        qt_dup += 1
                    else:
                        sheet[i][cab.index('Duplicidade')].value = 'não'
                except TypeError:
                    sheet[i][cab.index('Duplicidade')].value = ''

                sheet[i][cab.index('Códigos de Produto Analisados')].value = ', '.join(cliente.codigos_analisados)
                sheet[i][cab.index('Data Consulta')].value = today.strftime("%d/%m/%y")
    wb.save(plan)
    return qt_dup


def report_backup(clientes: list[DadosCliente]) -> None:
    """Gera um arquivo .txt de backup do report"""
    today = date.today()
    with open(f'report_duplicidade{today.strftime("%d.%m.%y")}.txt', 'w') as wb:
        for cliente in clientes:
            wb.write(cliente.__repr__() + '\n')


def criando_relatorio_csv_excel(clientes: list[DadosCliente]):
    """Gera um arquivo .csv e .xlsx com um relatório do que foi encontrado"""
    today = date.today()
    with open(f'relatorio_exec_int_{today.strftime("%d.%m.%y")}.csv', 'w', encoding='UTF8', newline='') as file:
        writer = csv.writer(file, dialect='excel')
        header = ['CNPJ', 'Cod_Cliente1', 'Regiao',  'T.O. 25kg', 'T.O. 50kg', 'Est. 40kg', 'Est. 50kg',
                  'códigos analisados']
        writer.writerow(header)
        for c in clientes:
            try:
                if c.cimento_todas_obras['25 to'] > 1 or c.cimento_todas_obras['50 to'] > 1 or \
                        c.cimento_estrutural['40 e'] > 1 or c.cimento_estrutural['50 e'] > 1:
                    data = [c.identificacao[0], c.identificacao[1], c.identificacao[2], c.cimento_todas_obras['25 to'],
                            c.cimento_todas_obras['50 to'], c.cimento_estrutural['40 e'], c.cimento_estrutural['50 e'],
                            ', '.join(c.codigos_analisados)]
                    writer.writerow(data)
            except KeyError:
                pass
    df = pd.read_csv(f'relatorio_exec_int_{today.strftime("%d.%m.%y")}.csv')
    df.to_excel(f'relatorio_exec_int_{today.strftime("%d.%m.%y")}.xlsx')


# usado para criar pastas com prints
def criacao_pastas():
    today = date.today()
    makedirs(f'\{today.strftime("%d.%m.%y")}', exist_ok=True)
    makedirs(f'\{today.strftime("%d.%m.%y")}\T.O.', exist_ok=True)
    makedirs(f'\{today.strftime("%d.%m.%y")}\Est.', exist_ok=True)


# necessário implementar
def enviar_email(qtd: int, emails: tuple, plan: str, dupli: int) -> str:
    today = date.today()
    relatorio = f"Prezados,\nA tarefa de check duplicidade de produtos na Loja virtual em " \
                f"{today.strftime('%d.%m.%y')} foi concluída.\nForam checadas {qtd} micro-regiões, das" \
                f" quais {dupli} estão com duplicidades. o check está registrado em {plan}\nAtt."
    return relatorio
